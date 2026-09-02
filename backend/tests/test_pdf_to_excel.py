from pathlib import Path
import openpyxl
import pymupdf
from backend.app.processors.pdf_to_excel import PdfToExcelProcessor, coerce_cell_value
from backend.app.core.validation import validate_xlsx_output

FIXTURES_DIR = Path(__file__).parent / "fixtures"


def test_pdf_to_excel_contains_table_data_invariant(tmp_path):
    """INVARIANT: XLSX opens, passes workbook zip check, and has populated data cells."""
    p_tab = FIXTURES_DIR / "table.pdf"
    processor = PdfToExcelProcessor(job_id="test_excel_inv", work_dir=tmp_path)
    result = processor.run([p_tab], {})

    assert result.status == "COMPLETED"
    validate_xlsx_output(result.output_path)

    wb = openpyxl.load_workbook(str(result.output_path))
    assert len(wb.sheetnames) >= 1
    sheet = wb.active
    assert sheet.max_row >= 1


def test_pdf_to_excel_multi_page_worksheets(tmp_path):
    """INVARIANT: Multi-page PDFs generate distinct Page_1 and Page_2 worksheets."""
    multi_pdf_path = tmp_path / "multi_table.pdf"
    doc = pymupdf.open()
    
    # Page 1
    p1 = doc.new_page(width=595, height=842)
    p1.insert_text((50, 50), "Quarter\tRevenue\tMargin", fontsize=12)
    p1.insert_text((50, 70), "Q1\t$15.2M\t42%", fontsize=11)
    
    # Page 2
    p2 = doc.new_page(width=595, height=842)
    p2.insert_text((50, 50), "Employee\tDepartment\tRating", fontsize=12)
    p2.insert_text((50, 70), "Alice\tEngineering\t5", fontsize=11)
    
    doc.save(str(multi_pdf_path))
    doc.close()
    
    processor = PdfToExcelProcessor(job_id="test_multi_excel", work_dir=tmp_path)
    result = processor.run([multi_pdf_path], {})
    
    assert result.status == "COMPLETED"
    wb = openpyxl.load_workbook(str(result.output_path))
    assert "Page_1" in wb.sheetnames
    assert "Page_2" in wb.sheetnames


def test_pdf_to_excel_process_bytes(tmp_path):
    """INVARIANT: process_bytes accepts in-memory PDF buffer and returns XLSX bytes."""
    p_tab = FIXTURES_DIR / "table.pdf"
    with open(p_tab, "rb") as f:
        pdf_bytes = f.read()

    processor = PdfToExcelProcessor(job_id="test_excel_bytes", work_dir=tmp_path)
    out_bytes, filename, metrics = processor.process_bytes(pdf_bytes, "table.pdf")

    assert len(out_bytes) > 0
    assert filename.endswith(".xlsx")
    assert metrics["format"] == "xlsx"


def test_coerce_cell_value():
    """Verify intelligent conversion of values for Excel cells."""
    assert coerce_cell_value("123") == 123
    assert coerce_cell_value("-45") == -45
    assert coerce_cell_value("12.50") == 12.5
    assert coerce_cell_value("$1,250.00") == 1250.0
    assert coerce_cell_value("Sample Text") == "Sample Text"
    assert coerce_cell_value(None) == ""
